from openpyxl.workbook import Workbook
from openpyxl import load_workbook

# Create a workbook object
#wb = Workbook()

# Create an active worksheet
#ws = wb.active

# Load existing worksheet
wb = load_workbook('StatsHunters-Export-2022-11-06.xlsx')
ws = wb.active

# # Set a variable
# name = ws['A2'].value
#
# # Print something from our spreadsheet
# print(name) #Since Python is a OO language, without value is just prints the object and not its value
#
# # Grab column A
# column_a = ws['A']
# print(column_a) #Does not work, again, because it is printing the tuple. Run a look to extract each one, one-by-one
#
# for i in column_a:
#     print(i.value)
#
# # Grab a range
# range = ws['A2:C10']
# for i in range:
#     for j in i:
#         print(j.value)

# # Iterate through rows
# for row in ws.iter_rows(min_row = 2, max_row = 10, min_col = 1, max_col = 3, values_only = True):
#     for cell in row:
#         print(cell)
#
# # Iterate through cols
# for row in ws.iter_cols(min_row = 2, max_row = 10, min_col = 1, max_col = 3, values_only = True):
#     for cell in col:
#         print(cell)

# # Change a cell
# ws["J2"] = "FS"
#
# # Save an excel spreadsheet
# wb.save('StatsHunters-Export-2022-11-06.xlsx')
# print(' File was saved...')

# # Change many cells
# km = [10, 15, 21]
#
# starting_row = 2
#
# for km in km:
#     ws.cell(row = starting_row, column = 11).value = km
#     starting_row += 1


