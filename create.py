from openpyxl.workbook import Workbook
from openpyxl import load_workbook

# Create a workbook object
wb = Workbook()

# Create an active worksheet
ws = wb.active

# Create worksheet title
ws.title = "Names and colors"

# Create a Python List of names
names = ["Dan", "April", "Neal", "Sara"]
colors = ["Blue", "Black", "Red", "Green"]
nums = [10, 8, 9, 12]

ws["A1"] = "Names"
ws["B1"] = "Colors"
ws["C1"] = "Favorite number"

# Add names to ws
starting_row = 2
for name in names:
    ws.cell(row = starting_row, column = 1). value = name
    starting_row += 1

starting_row = 2
for color in colors:
    ws.cell(row = starting_row, column = 2). value = color
    starting_row += 1

starting_row = 2
for num in nums:
    ws.cell(row = starting_row, column = 3). value = num
    starting_row += 1

#Use a formula
ws["C6"] = "=sum(C2:C5)"

# Save an excel spreadsheet
wb.save('names.xlsx')
print(' File was saved...')

