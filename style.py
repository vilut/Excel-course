from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side

# Create a workbook object
wb = load_workbook('names.xlsx')

# Create an active worksheet
ws = wb.active

# Select cell
cell = ws['A1']
cell.font = Font(
    size=30,
    bold = True,
    italic = False,
    color = "253bb8")

# Define a border
my_bd = Side(style="thick", color="000000")

B3 = ws['B3']
B3.border = Border(
    left = my_bd,
    right = my_bd,
    top = my_bd,
    bottom = my_bd)

# Save an excel spreadsheet
wb.save('names2.xlsx')
print(' File was saved...')